"""
Export utilities for generating documentation
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from typing import IO
from ..models.diagram import Diagram


class ExcelExporter:
    """Export diagram to Excel format"""

    def __init__(self, diagram: Diagram):
        self.diagram = diagram

    def export(self, output_path: str):
        """
        Export diagram to Excel file.

        Args:
            output_path: Path to save Excel file
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        self._create_overview_sheet(wb)
        self._create_tables_list_sheet(wb)
        self._create_table_detail_sheets(wb)

        # Save workbook
        wb.save(output_path)

    def _create_overview_sheet(self, wb: Workbook):
        """Create overview sheet"""
        ws = wb.create_sheet("Overview")

        # Header style
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)

        # Title
        ws['A1'] = "ERD Overview"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill

        # Diagram info
        ws['A3'] = "Diagram Name:"
        ws['B3'] = self.diagram.name
        ws['A4'] = "Database Type:"
        ws['B4'] = self.diagram.database_type
        ws['A5'] = "Version:"
        ws['B5'] = self.diagram.version
        ws['A6'] = "Description:"
        ws['B6'] = self.diagram.description or ""

        # Statistics
        stats = self.diagram.get_statistics()
        ws['A8'] = "Statistics"
        ws['A8'].font = Font(bold=True)

        ws['A9'] = "Total Tables:"
        ws['B9'] = stats['total_tables']
        ws['A10'] = "Total Columns:"
        ws['B10'] = stats['total_columns']
        ws['A11'] = "Total Relationships:"
        ws['B11'] = stats['total_relationships']
        ws['A12'] = "Total Primary Keys:"
        ws['B12'] = stats['total_primary_keys']
        ws['A13'] = "Total Foreign Keys:"
        ws['B13'] = stats['total_foreign_keys']

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def _create_tables_list_sheet(self, wb: Workbook):
        """Create tables list sheet"""
        ws = wb.create_sheet("Tables List")

        # Headers
        headers = ["Table Name", "Logical Name", "Columns Count", "Comment"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Table data
        for row, table in enumerate(self.diagram.tables, start=2):
            ws.cell(row=row, column=1, value=table.physical_name)
            ws.cell(row=row, column=2, value=table.logical_name or "")
            ws.cell(row=row, column=3, value=len(table.columns))
            ws.cell(row=row, column=4, value=table.comment or "")

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50

    def _create_table_detail_sheets(self, wb: Workbook):
        """Create detailed sheet for each table"""
        for table in self.diagram.tables:
            # Sheet name (Excel limit: 31 chars)
            sheet_name = table.physical_name[:31]
            ws = wb.create_sheet(sheet_name)

            # Table info
            ws['A1'] = "Table:"
            ws['B1'] = table.physical_name
            ws['B1'].font = Font(bold=True, size=12)

            ws['A2'] = "Logical Name:"
            ws['B2'] = table.logical_name or ""

            ws['A3'] = "Comment:"
            ws['B3'] = table.comment or ""

            # Column headers
            col_headers = ["Column Name", "Logical Name", "Type", "Length", "Nullable", "PK", "FK", "Default", "Comment"]
            for col, header in enumerate(col_headers, start=1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Column data
            for row, column in enumerate(table.columns, start=6):
                ws.cell(row=row, column=1, value=column.physical_name)
                ws.cell(row=row, column=2, value=column.logical_name or "")
                ws.cell(row=row, column=3, value=column.data_type.upper())
                ws.cell(row=row, column=4, value=column.length or column.precision or "")
                ws.cell(row=row, column=5, value="YES" if column.nullable else "NO")
                ws.cell(row=row, column=6, value="✓" if column.primary_key else "")
                ws.cell(row=row, column=7, value="✓" if column.is_foreign_key() else "")
                ws.cell(row=row, column=8, value=str(column.default_value) if column.default_value is not None else "")
                ws.cell(row=row, column=9, value=column.comment or "")

            # Column widths
            for col in range(1, 10):
                ws.column_dimensions[chr(64 + col)].width = 15


class HTMLExporter:
    """Export diagram to HTML format"""

    def __init__(self, diagram: Diagram):
        self.diagram = diagram

    def export(self, output_path: str):
        """
        Export diagram to HTML file.

        Args:
            output_path: Path to save HTML file
        """
        html = self._generate_html()

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)

    def _generate_html(self) -> str:
        """Generate HTML content"""
        stats = self.diagram.get_statistics()

        html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.diagram.name} - ERD Documentation</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            line-height: 1.6;
            color: #333;
            background: #f5f5f5;
            padding: 20px;
        }}

        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}

        h1 {{
            color: #2C3E50;
            border-bottom: 3px solid #4A90E2;
            padding-bottom: 10px;
            margin-bottom: 30px;
        }}

        h2 {{
            color: #4A90E2;
            margin-top: 40px;
            margin-bottom: 20px;
            border-bottom: 2px solid #ECF0F1;
            padding-bottom: 10px;
        }}

        .info-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 40px;
        }}

        .info-card {{
            background: #ECF0F1;
            padding: 20px;
            border-radius: 6px;
        }}

        .info-card label {{
            font-weight: bold;
            color: #7F8C8D;
            display: block;
            margin-bottom: 5px;
        }}

        .info-card value {{
            font-size: 1.2em;
            color: #2C3E50;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
        }}

        th {{
            background: #4A90E2;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}

        td {{
            padding: 10px 12px;
            border-bottom: 1px solid #ECF0F1;
        }}

        tr:hover {{
            background: #F8F9FA;
        }}

        .table-detail {{
            background: #F8F9FA;
            padding: 20px;
            border-radius: 6px;
            margin-bottom: 30px;
        }}

        .table-name {{
            font-size: 1.4em;
            color: #2C3E50;
            font-weight: bold;
            margin-bottom: 10px;
        }}

        .table-comment {{
            color: #7F8C8D;
            font-style: italic;
            margin-bottom: 15px;
        }}

        .pk {{
            color: #F39C12;
            font-weight: bold;
        }}

        .fk {{
            color: #9B59B6;
            font-weight: bold;
        }}

        @media print {{
            body {{
                background: white;
                padding: 0;
            }}
            .container {{
                box-shadow: none;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 {self.diagram.name}</h1>

        <div class="info-grid">
            <div class="info-card">
                <label>Database Type</label>
                <value>{self.diagram.database_type.upper()}</value>
            </div>
            <div class="info-card">
                <label>Version</label>
                <value>{self.diagram.version}</value>
            </div>
            <div class="info-card">
                <label>Total Tables</label>
                <value>{stats['total_tables']}</value>
            </div>
            <div class="info-card">
                <label>Total Columns</label>
                <value>{stats['total_columns']}</value>
            </div>
            <div class="info-card">
                <label>Relationships</label>
                <value>{stats['total_relationships']}</value>
            </div>
            <div class="info-card">
                <label>Last Updated</label>
                <value>{self.diagram.updated_at.strftime('%Y-%m-%d %H:%M')}</value>
            </div>
        </div>

        <h2>📋 Tables Overview</h2>
        <table>
            <thead>
                <tr>
                    <th>Table Name</th>
                    <th>Logical Name</th>
                    <th>Columns</th>
                    <th>Description</th>
                </tr>
            </thead>
            <tbody>
"""

        # Tables list
        for table in self.diagram.tables:
            html += f"""
                <tr>
                    <td><a href="#table-{table.physical_name}">{table.physical_name}</a></td>
                    <td>{table.logical_name or '-'}</td>
                    <td>{len(table.columns)}</td>
                    <td>{table.comment or '-'}</td>
                </tr>
"""

        html += """
            </tbody>
        </table>

        <h2>📑 Table Details</h2>
"""

        # Table details
        for table in self.diagram.tables:
            html += f"""
        <div class="table-detail" id="table-{table.physical_name}">
            <div class="table-name">{table.physical_name}</div>
            {f'<div class="table-comment">{table.comment}</div>' if table.comment else ''}

            <table>
                <thead>
                    <tr>
                        <th>Column</th>
                        <th>Logical Name</th>
                        <th>Type</th>
                        <th>Null</th>
                        <th>Key</th>
                        <th>Default</th>
                        <th>Comment</th>
                    </tr>
                </thead>
                <tbody>
"""

            for column in table.columns:
                key_info = ""
                if column.primary_key:
                    key_info = '<span class="pk">PK</span>'
                elif column.is_foreign_key():
                    key_info = '<span class="fk">FK</span>'

                html += f"""
                    <tr>
                        <td><strong>{column.physical_name}</strong></td>
                        <td>{column.logical_name or '-'}</td>
                        <td>{column.get_full_type()}</td>
                        <td>{'YES' if column.nullable else 'NO'}</td>
                        <td>{key_info}</td>
                        <td>{column.default_value if column.default_value is not None else '-'}</td>
                        <td>{column.comment or '-'}</td>
                    </tr>
"""

            html += """
                </tbody>
            </table>
        </div>
"""

        html += """
    </div>
</body>
</html>
"""

        return html
